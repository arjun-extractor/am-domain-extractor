import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Page Configuration
st.set_page_config(
    page_title="AM Domain Extractor - Multi-Threaded High-Speed Edition",
    page_icon="⚡",
    layout="wide"
)

# Custom Styling to match the professional dashboard
st.markdown("""
<style>
    .main-title {
        font-size: 1.8rem;
        font-weight: 700;
        color: #111827;
        margin-bottom: 0px;
    }
    .sub-title {
        font-size: 0.85rem;
        color: #6B7280;
        margin-bottom: 15px;
    }
    .metric-box {
        background-color: #1E3A8A;
        color: white;
        padding: 10px;
        border-radius: 6px;
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.markdown("### ⚡ AM Domain Extractor")
    st.caption("Multi-Threaded High-Speed Edition")
    st.markdown("---")
    if st.button("📊 Dashboard", use_container_width=True, type="primary"):
        pass
    if st.button("📥 Download Results ⬇️", use_container_width=True):
        pass
    
    st.markdown("---")
    st.markdown("<div class='metric-box'><b>PROCESSED QUERIES</b><br><span style='font-size: 1.4rem;'>88 / 250</span></div>", unsafe_allow_html=True)
    st.markdown("<div class='metric-box'><b>COMPLETED PERCENT</b><br><span style='font-size: 1.4rem;'>35%</span></div>", unsafe_allow_html=True)
    st.markdown("<div class='metric-box'><b>DIRECTORIES / SKIPPED</b><br><span style='font-size: 1.4rem;'>8</span></div>", unsafe_allow_html=True)
    st.markdown("<div class='metric-box'><b>DOMAINS FOUND</b><br><span style='font-size: 1.4rem;'>2077</span></div>", unsafe_allow_html=True)

# Main Content Area
st.markdown('<div class="main-title">A Powerful Lead Generation Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Multi-threaded Google Maps domain extraction platform</div>', unsafe_allow_html=True)

col1, col2 = st.columns([1.2, 1])

with col1:
    st.markdown("<b>1. Paste Keywords (One per line):</b>", unsafe_allow_html=True)
    keywords = st.text_area(
        "Keywords", 
        value="Plumbers\nElectricians\nHouse Cleaners\nCarpet Cleaners", 
        height=120,
        label_visibility="collapsed"
    )

with col2:
    st.markdown("<b>2. Target Locations:</b>", unsafe_allow_html=True)
    locations = st.multiselect(
        "Locations",
        ["United States", "California", "New York", "San Diego", "Los Angeles", "San Francisco"],
        default=["San Diego, California", "Los Angeles, California", "San Francisco, California"],
        label_visibility="collapsed"
    )
    st.caption(f"Selected: {len(locations)} Locations")
    
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        threads = st.slider("Threads (1-30)", 1, 30, 10)
    with col_t2:
        max_results = st.selectbox("Max Results:", ["Unlimited", "500", "1000", "5000"])

st.markdown("---")

# Control Buttons Bar
b_col1, b_col2, b_col3, b_col4, b_col5 = st.columns([1, 1, 1, 1, 2])
with b_col1:
    start_btn = st.button("▶️ Start", type="primary", use_container_width=True)
with b_col2:
    stop_btn = st.button("⏹️ Stop", use_container_width=True)
with b_col3:
    pause_btn = st.button("⏸️ Pause", use_container_width=True)
with b_col4:
    reset_btn = st.button("🔄 Reset", use_container_width=True)

st.markdown("---")
st.markdown("### Live Multi-Threaded Data Stream")

data = {
    "Search Query": [
        "Air Conditioning Repair in San Diego, California",
        "Air Conditioning Repair in San Diego, California",
        "Appliance Repair Services in Los Angeles, California",
        "Appliance Repair Services in Los Angeles, California",
        "Appliance Repair Services in San Francisco, California",
        "Appliance Repair Services in San Francisco, California"
    ],
    "Business Domain": [
        "hvac-sdpro.com",
        "bigcityhomeservice.com",
        "appliance-repair.ai",
        "prme-fix.com",
        "appliance-insight.com",
        "allstateappliancerepair.com"
    ],
    "Full URL": [
        "http://hvac-sdpro.com/",
        "https://bigcityhomeservice.com/?utm_campaign=gmb",
        "https://appliance-repair.ai/",
        "http://prme-fix.com/",
        "https://appliance-insight.com/",
        "https://www.allstateappliancerepair.com"
    ]
}

df_live = pd.DataFrame(data)
st.dataframe(df_live, use_container_width=True, height=250)

# Excel Export Helper
def generate_excel(df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Leads"
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'), top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB'))
    
    headers = list(df.columns)
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        for c_idx in range(1, len(row) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)
        
    ws.freeze_panes = "A2"
    wb.save(output)
    output.seek(0)
    return output

excel_data = generate_excel(df_live)
st.download_button(
    label="📥 Download Extracted Leads as Excel (.xlsx)",
    data=excel_data,
    file_name="AM_Domain_Extractor_Results.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
